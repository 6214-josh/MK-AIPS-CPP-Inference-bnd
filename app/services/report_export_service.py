from pathlib import Path
from datetime import datetime
import csv
import html
from app.core.database import fetch_all, execute_returning_id
from app.core.schema_guard import ensure_extra_schema

REPORT_DIR = Path(__file__).resolve().parents[2] / "generated_reports"
REPORT_DIR.mkdir(parents=True, exist_ok=True)

REPORT_COLUMNS = [
    ("cnc_machine_id", "CNC 機台"),
    ("machine_status", "機台狀態"),
    ("power_kw", "即時功率 kW"),
    ("demand_kw", "最大需量 kW"),
    ("thd_current", "THD 電流"),
    ("power_factor", "功率因數"),
    ("power_kwh", "累積用電 kWh"),
    ("estimated_machine_status", "AI 判斷狀態"),
    ("machine_abnormal_power_flag", "電力異常"),
]

def _rows():
    ensure_extra_schema()
    rows = fetch_all("""
        SELECT
            COALESCE(r.cnc_machine_id, l.cnc_machine_id) AS cnc_machine_id,
            COALESCE(s.machine_status, f.estimated_machine_status, 'UNKNOWN') AS machine_status,
            r.power_kw,
            r.demand_kw,
            r.thd_current,
            r.power_factor,
            r.power_kwh,
            f.estimated_machine_status,
            f.machine_abnormal_power_flag,
            r.collect_time
        FROM aips_electric_cnc_link l
        LEFT JOIN LATERAL (
            SELECT * FROM cnc_meter_raw_data r
            WHERE r.cnc_machine_id = l.cnc_machine_id
            ORDER BY r.meter_data_id DESC
            LIMIT 1
        ) r ON TRUE
        LEFT JOIN LATERAL (
            SELECT * FROM cnc_meter_feature f
            WHERE f.cnc_machine_id = l.cnc_machine_id
            ORDER BY f.feature_id DESC
            LIMIT 1
        ) f ON TRUE
        LEFT JOIN aips_sim_cnc_smart_meter s ON s.cnc_machine_id = l.cnc_machine_id
        ORDER BY l.cnc_machine_id
    """)
    if not rows:
        rows = fetch_all("""
            SELECT
                cnc_machine_id,
                'UNKNOWN' AS machine_status,
                power_kw,
                demand_kw,
                thd_current,
                power_factor,
                power_kwh,
                NULL AS estimated_machine_status,
                FALSE AS machine_abnormal_power_flag,
                collect_time
            FROM cnc_meter_raw_data
            ORDER BY meter_data_id DESC
            LIMIT 20
        """)
    return rows

def _summary(rows):
    n = len(rows) or 1
    total_kwh = sum(float(r.get("power_kwh") or 0) for r in rows)
    avg_power = sum(float(r.get("power_kw") or 0) for r in rows) / n
    max_demand = max([float(r.get("demand_kw") or 0) for r in rows] or [0])
    abnormal = sum(1 for r in rows if r.get("machine_abnormal_power_flag") or float(r.get("thd_current") or 0) >= 15)
    return {
        "total_kwh": round(total_kwh, 2),
        "avg_power": round(avg_power, 2),
        "max_demand": round(max_demand, 2),
        "abnormal": abnormal,
        "carbon": round(total_kwh * 0.494, 2),
    }

def _save_job(report_type: str, report_name: str, fmt: str, path: Path):
    return execute_returning_id("""
        INSERT INTO aips_report_job (report_type, report_name, file_format, file_path, job_status, created_by)
        VALUES (%s, %s, %s, %s, 'DONE', 'admin')
        RETURNING report_job_id
    """, (report_type, report_name, fmt, str(path)), "report_job_id")

def generate_excel_report():
    ensure_extra_schema()
    rows = _rows()
    s = _summary(rows)
    now = datetime.now()
    filename = f"aips_oee_energy_report_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = REPORT_DIR / filename

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()
    ws = wb.active
    ws.title = "AIPS OEE Energy"

    blue = "1D4ED8"
    light_blue = "DBEAFE"
    dark = "0F172A"
    gray = "E5E7EB"
    danger = "FEE2E2"
    white = "FFFFFF"

    ws.merge_cells("A1:I1")
    ws["A1"] = "MinKing AIPS 智慧排程報表 - CNC 智慧電表 / OEE / DQN"
    ws["A1"].font = Font(size=18, bold=True, color=white)
    ws["A1"].fill = PatternFill("solid", fgColor=blue)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:I2")
    ws["A2"] = f"產生時間：{now.strftime('%Y-%m-%d %H:%M:%S')}｜資料來源：智慧電表、CNC 狀態、AIPS 特徵工程"
    ws["A2"].font = Font(size=11, color=dark)
    ws["A2"].fill = PatternFill("solid", fgColor="F8FAFC")

    cards = [
        ("總用電 kWh", s["total_kwh"]),
        ("平均功率 kW", s["avg_power"]),
        ("最大需量 kW", s["max_demand"]),
        ("異常機台數", s["abnormal"]),
        ("碳排 kg", s["carbon"]),
    ]
    start_col = 1
    for idx, (label, value) in enumerate(cards):
        col = start_col + idx * 2
        ws.cell(row=4, column=col, value=label)
        ws.cell(row=5, column=col, value=value)
        ws.cell(row=4, column=col).font = Font(bold=True, color=white)
        ws.cell(row=4, column=col).fill = PatternFill("solid", fgColor=blue)
        ws.cell(row=5, column=col).font = Font(size=16, bold=True, color=dark)
        ws.cell(row=5, column=col).fill = PatternFill("solid", fgColor=light_blue)
        ws.cell(row=4, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=5, column=col).alignment = Alignment(horizontal="center")

    header_row = 8
    for c, (_, label) in enumerate(REPORT_COLUMNS, 1):
        cell = ws.cell(row=header_row, column=c, value=label)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color=gray)
    for r_idx, row in enumerate(rows, header_row + 1):
        for c_idx, (key, _) in enumerate(REPORT_COLUMNS, 1):
            value = row.get(key)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if key in ("machine_abnormal_power_flag", "thd_current") and (row.get("machine_abnormal_power_flag") or float(row.get("thd_current") or 0) >= 15):
                cell.fill = PatternFill("solid", fgColor=danger)

    for col in range(1, len(REPORT_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.freeze_panes = "A9"
    ws.auto_filter.ref = f"A8:I{max(header_row+1, header_row+len(rows))}"

    # Chart
    if rows:
        chart = BarChart()
        chart.title = "CNC 即時功率 kW"
        chart.y_axis.title = "kW"
        chart.x_axis.title = "CNC"
        data = Reference(ws, min_col=3, min_row=8, max_row=8+len(rows))
        cats = Reference(ws, min_col=1, min_row=9, max_row=8+len(rows))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 15
        ws.add_chart(chart, "K8")

    wb.save(path)
    job_id = _save_job("OEE_ENERGY", "每日 OEE / 智慧電表報表", "Excel", path)
    return {"report_job_id": job_id, "file": path, "filename": filename}

def generate_csv_report():
    ensure_extra_schema()
    rows = _rows()
    now = datetime.now()
    filename = f"aips_oee_energy_report_{now.strftime('%Y%m%d_%H%M%S')}.csv"
    path = REPORT_DIR / filename
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([label for _, label in REPORT_COLUMNS])
        for row in rows:
            writer.writerow([row.get(key, "") for key, _ in REPORT_COLUMNS])
    job_id = _save_job("OEE_ENERGY", "每日 OEE / 智慧電表 CSV", "CSV", path)
    return {"report_job_id": job_id, "file": path, "filename": filename}

def generate_bi_html_report():
    ensure_extra_schema()
    rows = _rows()
    s = _summary(rows)
    now = datetime.now()
    filename = f"aips_bi_dashboard_{now.strftime('%Y%m%d_%H%M%S')}.html"
    path = REPORT_DIR / filename
    trs = []
    for r in rows:
        abnormal = r.get("machine_abnormal_power_flag") or float(r.get("thd_current") or 0) >= 15
        cls = "danger" if abnormal else ""
        tds = "".join(f"<td>{html.escape(str(r.get(k, '') if r.get(k) is not None else ''))}</td>" for k, _ in REPORT_COLUMNS)
        trs.append(f"<tr class='{cls}'>{tds}</tr>")

    path.write_text(f"""
<!doctype html>
<html lang="zh-Hant">
<head>
<meta charset="utf-8">
<title>MinKing AIPS BI Dashboard</title>
<style>
body{{font-family:Arial,'Microsoft JhengHei',sans-serif;background:#f4f7fb;color:#0f172a;margin:0;padding:28px}}
.header{{background:linear-gradient(135deg,#1d4ed8,#0f172a);color:white;border-radius:20px;padding:24px;margin-bottom:18px}}
.cards{{display:grid;grid-template-columns:repeat(5,1fr);gap:14px;margin-bottom:18px}}
.card{{background:white;border-radius:16px;padding:18px;box-shadow:0 8px 24px rgba(15,23,42,.08)}}
.label{{color:#64748b;font-size:13px}} .value{{font-size:28px;font-weight:900;margin-top:8px}}
table{{width:100%;border-collapse:collapse;background:white;border-radius:16px;overflow:hidden}}
th{{background:#0f172a;color:white;padding:12px}}td{{padding:10px;border-bottom:1px solid #e5e7eb;text-align:center}}
tr.danger td{{background:#fee2e2}}
</style>
</head>
<body>
<div class="header"><h1>MinKing AIPS BI Dashboard</h1><p>產生時間：{now.strftime('%Y-%m-%d %H:%M:%S')}｜CNC 智慧電表 / OEE / DQN 報表</p></div>
<div class="cards">
<div class="card"><div class="label">總用電 kWh</div><div class="value">{s['total_kwh']}</div></div>
<div class="card"><div class="label">平均功率 kW</div><div class="value">{s['avg_power']}</div></div>
<div class="card"><div class="label">最大需量 kW</div><div class="value">{s['max_demand']}</div></div>
<div class="card"><div class="label">異常機台數</div><div class="value">{s['abnormal']}</div></div>
<div class="card"><div class="label">碳排 kg</div><div class="value">{s['carbon']}</div></div>
</div>
<table>
<thead><tr>{''.join(f'<th>{label}</th>' for _, label in REPORT_COLUMNS)}</tr></thead>
<tbody>{''.join(trs)}</tbody>
</table>
</body>
</html>
""", encoding="utf-8")
    job_id = _save_job("BI_DASHBOARD", "AIPS BI Dashboard", "HTML", path)
    return {"report_job_id": job_id, "file": path, "filename": filename}

def generate_pdf_report():
    ensure_extra_schema()
    rows = _rows()
    s = _summary(rows)
    now = datetime.now()
    filename = f"aips_oee_energy_report_{now.strftime('%Y%m%d_%H%M%S')}.pdf"
    path = REPORT_DIR / filename

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    styles = getSampleStyleSheet()
    styles["Title"].fontName = "STSong-Light"
    styles["Normal"].fontName = "STSong-Light"

    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    story = []
    story.append(Paragraph("MinKing AIPS 智慧排程報表 - CNC 智慧電表 / OEE / DQN", styles["Title"]))
    story.append(Paragraph(f"產生時間：{now.strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]))
    story.append(Spacer(1, 12))

    summary_data = [
        ["總用電 kWh", "平均功率 kW", "最大需量 kW", "異常機台數", "碳排 kg"],
        [s["total_kwh"], s["avg_power"], s["max_demand"], s["abnormal"], s["carbon"]],
    ]
    summary_table = Table(summary_data, colWidths=[110]*5)
    summary_table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), "STSong-Light"),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1D4ED8")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("BACKGROUND", (0,1), (-1,1), colors.HexColor("#DBEAFE")),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ("FONTSIZE", (0,0), (-1,-1), 10),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 14))

    table_data = [[label for _, label in REPORT_COLUMNS]]
    for r in rows:
        table_data.append([str(r.get(k, "") if r.get(k) is not None else "") for k, _ in REPORT_COLUMNS])
    table = Table(table_data, repeatRows=1, colWidths=[85,75,70,70,60,60,75,80,65])
    style = TableStyle([
        ("FONTNAME", (0,0), (-1,-1), "STSong-Light"),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#CBD5E1")),
        ("FONTSIZE", (0,0), (-1,-1), 8),
    ])
    for idx, r in enumerate(rows, start=1):
        if r.get("machine_abnormal_power_flag") or float(r.get("thd_current") or 0) >= 15:
            style.add("BACKGROUND", (0,idx), (-1,idx), colors.HexColor("#FEE2E2"))
    table.setStyle(style)
    story.append(table)
    doc.build(story)

    job_id = _save_job("OEE_ENERGY", "每日 OEE / 智慧電表 PDF", "PDF", path)
    return {"report_job_id": job_id, "file": path, "filename": filename}

def generate_report(format_name: str):
    fmt = (format_name or "excel").lower()
    if fmt in ("excel", "xlsx"):
        return generate_excel_report()
    if fmt == "pdf":
        return generate_pdf_report()
    if fmt in ("bi", "html"):
        return generate_bi_html_report()
    if fmt == "csv":
        return generate_csv_report()
    raise ValueError(f"不支援的報表格式：{format_name}")
